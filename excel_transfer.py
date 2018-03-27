#!/usr/bin/python
# -*- coding: utf-8 -*-

import sys
import os
import openpyxl

# excel column setting
COLUMN_COUNT_CONFIG = 13

# input column name setting
COLUMN_NAME_LIST = ['家庭人数', '保费', 
                    '本人姓名', '本人身份证号码', 
                    '亲属1姓名', '亲属1身份证号码',
                    '亲属2姓名', '亲属2身份证号码',
                    '亲属3姓名', '亲属3身份证号码',
                    '亲属4姓名', '亲属4身份证号码']
OUTPUT_HEADER = ['家庭人数', '姓名', '身份证号码', '本人/亲属', '人均保费']


# TODO id check: unique/legal/age
def check_id(id):
    return 0

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print "python excel_transfer.py input_dir output_dir"
        sys.exit(-1)

input_dir = sys.argv[1]
output_dir =sys.argv[2]
input_file_list = os.listdir(input_dir);

# check input exist
if not os.path.isdir(input_dir):
    print "input dir:" + input_dir + " not exist!"
    sys.exit(-1)

# remove out dir
if os.path.isdir(output_dir):
    print "rm exist output dit:" + output_dir
    os.system("rm -rf %s" % output_dir)
os.system("mkdir %s" % output_dir)

# read each input excel
for each_input_file in input_file_list:
    # get excel sheet
    file_name = "%s/%s" % (input_dir, each_input_file)
    file_excel = openpyxl.load_workbook(file_name)
    sheet = file_excel.get_sheet_by_name(file_excel.get_sheet_names()[0])

    # get sheet count
    family_count = sheet.max_row
    family_attr_count = sheet.max_column
    if family_attr_count != COLUMN_COUNT_CONFIG:
        print "ERROR: file-%s has %d column, not equal expect %d" % (file_name, family_attr_count, COLUMN_COUNT_CONFIG)
        sys.exit(-1)
    
    print "INFO: file-%s has %d familys" % (file_name, family_count - 1)

    # get sheet header
    header = {}
    index = 1
    for each_value in list(sheet.rows)[0]:
        header[each_value.value.encode('utf8')] = index;
        index += 1;

    # check column name
    for each_column_name in COLUMN_NAME_LIST:
        if each_column_name not in header:
            print "ERROR: file-%s don't have column-%s" % (file_name, each_column_name)
            sys.exit(-1)
        # print "%s %d" % (each_column_name, header[each_column_name])

    # new output file
    output_excel = openpyxl.Workbook()
    output_sheet = output_excel.get_active_sheet();
    output_sheet.append(OUTPUT_HEADER);
    # split each family
    for xrow in xrange(2, family_count):
        member_count = int(sheet.cell(row = xrow, column = header["家庭人数"]).value)
        member_cost = float(sheet.cell(row = xrow, column = header["保费"]).value) / float(member_count)
        if member_count < 1 or member_count > 5:
            print "ERROR: file-%s line-%d has %d member"
            output_excel.save(filename = "%s/split_%s" % (output_dir, each_input_file))
            sys.exit(-1)
        row_name = sheet.cell(row = xrow, column = header["本人姓名"]).value.encode('utf8')
        row_id =sheet.cell(row = xrow, column = header["本人身份证号码"]).value.encode('utf8')
        if check_id(row_id) != 0:
            sys.exit(-1)
        row_charc = '本人'

        # insert self row
        row = [member_count, row_name, row_id, row_charc, member_cost]
        output_sheet.append(row)

        # insert relate row
        for relate_i in xrange(1, member_count):
            row_name = sheet.cell(row = xrow, column = header["亲属%d姓名" % relate_i]).value.encode('utf8')
            row_id =sheet.cell(row = xrow, column = header["亲属%d身份证号码" % relate_i]).value.encode('utf8')
            if check_id(row_id) != 0:
                sys.exit(-1)
            row_charc = '亲属'
            row = [member_count, row_name, row_id, row_charc, member_cost]
            output_sheet.append(row)

    # save output file
    output_excel.save(filename = "%s/split_%s" % (output_dir, each_input_file))
