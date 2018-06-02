#!/usr/bin/python
# -*- coding: utf-8 -*-

import sys
import os
import openpyxl

# input file type
FILE_TYPE = ['失独', '伤残', '低保']

# input sheet type
SHEET_TYPE = [u'1人', u'2人', u'3人', u'4人']

# output header
OUTPUT_HEADER = ['家庭人数', '本人/亲属', '姓名', '身份证号码', '人均保费', '街道', '类别']

# output column
OUTPUT_FIELD = ['A', 'D', 'E', 'F', 'P']

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print "python excel_transfer.py input_dir output_file"
        sys.exit(-1)

    input_dir = sys.argv[1]
    output_file =sys.argv[2]

    # check input exist
    if not os.path.isdir(input_dir):
        print "input dir:" + input_dir + " not exist!"
        sys.exit(-1)

    # remove outfile
    os.system("rm -rf %s" % output_file)
    
    output_excel = openpyxl.Workbook()
    output_sheet = output_excel.get_active_sheet()
    output_sheet.append(OUTPUT_HEADER)

    # get each input area
    input_area_list = os.listdir(input_dir);

    print "INFO:AREA LIST"
    for each_area in input_area_list:
        print each_area

    # get area's each input file 
    for each_area in input_area_list:
        # make output file
        area_input_file_list = os.listdir("%s/%s" % (input_dir, each_area))
        print "INFO: AREA %s LIST" % each_area
        for each_input_file in area_input_file_list:
            print each_input_file

        for each_input_file in area_input_file_list:
            excel_file_name = "%s/%s/%s" % (input_dir, each_area, each_input_file)
            # check file type
            file_type = "null"
            for each_type in FILE_TYPE:
                if each_type in each_input_file:
                    file_type = each_type
                    break
                
            if file_type == "null":
                print "WARNING: %s doesn't recognize file type %s" % (excel_file_name, '/'.join(FILE_TYPE))
                continue
            print "INFO: FILE TYPE = %s of %s" % (file_type, each_input_file)
            
            # get excel sheel list
            file_excel = openpyxl.load_workbook(excel_file_name, data_only = True)
            sheet_name_list = file_excel.get_sheet_names()
            print "INFO: %s sheet lists" % each_input_file
            for each_sheet_name in sheet_name_list:
                print each_sheet_name

            for each_sheet_type in SHEET_TYPE:
                if each_sheet_type not in sheet_name_list:
                    print "WARNING: %s don't has sheet %s" % (excel_file_name, each_sheet_type)
                    continue

                # get excel sheet 
                sheet = file_excel.get_sheet_by_name(each_sheet_type)
                print "INFO: %s sheet %s with %d row %d column" % (excel_file_name, 
                    each_sheet_type.encode('utf8'), sheet.max_row, sheet.max_column)
               
                for x_row in xrange(2, sheet.max_row):
                    # licence id 
                    licence = sheet.cell(row = x_row, column = ord('F') - ord('A') + 1).value
                    if licence == None:
                        continue
                    new_row = []
                    for each_field in OUTPUT_FIELD:
                        if each_sheet_type == u'1人' and each_field == 'P':
                            each_field = 'C'
                        x_col = ord(each_field) - ord('A') + 1
                        new_row.append(sheet.cell(row = x_row, column = x_col).value)
                    new_row.append(each_area)
                    new_row.append(file_type)
                    output_sheet.append(new_row)

        output_excel.save(filename = output_file)
